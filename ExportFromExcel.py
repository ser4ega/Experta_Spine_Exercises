import sqlite3
from openpyxl import load_workbook

def refresh_bd_from_excel():
    # Подключение к базе данных
    conn = sqlite3.connect('exercises.db')
    cursor = conn.cursor()



    # Загрузка данных из Excel файла
    wb = load_workbook('exercises.xlsx')
    ws1 = wb['Описания упражнений']
    ws2 = wb['Упражнения']
    ws3 = wb['Болезни']
    ws4 = wb['Ограничения упражнений']
    ws5 = wb['Отдел позвоночника']
    ws6 = wb['Эффекты упражнений']
    ws7 = wb['Сводная_Болезни_Эффект']

    # ВНИМАНИЕ удаление базы данных
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
    tables = cursor.fetchall()
    for table_name in tables:
        if table_name[0] != 'sqlite_sequence':
            cursor.execute(f"DROP TABLE {table_name[0]}")

    # Создание таблицы упражнений
    cursor.execute('''CREATE TABLE IF NOT EXISTS exercises_rel_deseases
                    (
                    exercise_id INTEGER,
                    exercise_name TEXT,
                    desease_id INTEGER,
                    effect_id INTEGER,
                    spine_section_id INTEGER,
                    description TEXT,
                    FOREIGN KEY(exercise_id) REFERENCES exercises_attr(id),
                    FOREIGN KEY(effect_id) REFERENCES effects(id),
                    FOREIGN KEY(spine_section_id) REFERENCES spine_section(id),
                    FOREIGN KEY(desease_id) REFERENCES spine_deseases(id))''')
    # Цикл для обработки каждой группы строк с данными. Заполняем атблицы названия упражнений и описания упражнений
    row_name = 2
    
    for row_group in range(1, ws1.max_row, 10):
        name = ws2.cell(row=row_name, column=1).value
        id_ =  ws2.cell(row=row_name, column=3).value

        desease_ids = ws2.cell(row=row_name, column=2).value.split(', ')
        effect_ids = str(ws2.cell(row=row_name, column=4).value).split(', ')
        spine_section_ids = str(ws2.cell(row=row_name, column=5).value).split(', ')
        description = "\n".join(str(ws1.cell(row=i, column=1).value) for i in range(row_group, row_group+10))

        for desease_id in desease_ids:
            for effect_id in effect_ids:
                for spine_section_id in spine_section_ids:
                    cursor.execute("INSERT INTO exercises_rel_deseases (exercise_id,exercise_name, desease_id, effect_id, spine_section_id, description) VALUES (?, ?, ?, ?, ?, ?)", (id_,name, desease_id,effect_id,spine_section_id,description))

        row_name = row_name + 1

    # Создание таблицы секции позвоночника в базе данных
    cursor.execute('CREATE TABLE IF NOT EXISTS spine_section (id Integer PRIMARY KEY, section TEXT)')
    for row in ws5.iter_rows(min_row=2, values_only=True):
        id_, section = row
        cursor.execute('INSERT INTO spine_section VALUES (?, ?)', (id_, section))

    # Создание таблицы эффекты упражнений в базе данных
    cursor.execute('CREATE TABLE IF NOT EXISTS effects (id Integer PRIMARY KEY, effect TEXT)')
    for row in ws6.iter_rows(min_row=2, values_only=True):
        id_, effect = row
        cursor.execute('INSERT INTO effects VALUES (?, ?)', (id_, effect))

    # Создание таблицы аттрибутов и ограничений для упражнений
    # conn.commit()
    cursor.execute('''CREATE TABLE IF NOT EXISTS exercises_attr
                    (id INTEGER PRIMARY KEY,
                weight FLOAT,
                age INTEGER,
                height FLOAT,
                level TEXT,
                equipment TEXT,
                time TEXT)''')
    for row in ws4.iter_rows(min_row=2, values_only=True):
        
        exercise_id = row[0]
        weight = row[1]
        age = row[2]
        height = row[3]
        level = row[4]
        equipment = row[5]
        time = row[6]
        # Добавляем данные в таблицу exercises_attr
        cursor.execute("INSERT INTO exercises_attr (id, weight, age, height, level, equipment, time) VALUES (?, ?, ?, ?, ?, ?, ?)", 
                (exercise_id, weight, age, height, level, equipment, time))
    # conn.commit()



    # Создание таблицы заболеваний в базе данных
    cursor.execute('''CREATE TABLE IF NOT EXISTS spine_deseases 
                    (id Integer PRIMARY KEY,
                    desease TEXT, 
                    id_section TEXT, 
                    recommends TEXT,
                    FOREIGN KEY (id_section) REFERENCES spine_section(id)
                    )''')

    # Чтение данных из листа Excel-файла и запись их в базу данных
    for row in ws3.iter_rows(min_row=2, values_only=True):
        id_, desease, id_section, recommended_effect_id_string, recommends = row

        # recommended_effect_ids = str(recommended_effect_id_string).split(', ')
        # for recomend_effect in recommended_effect_ids:
        cursor.execute('''INSERT INTO spine_deseases (id,
                    desease, 
                    id_section,
                    recommends) VALUES (?, ?, ?, ?)''', (id_, desease, id_section,  recommends))
    # создание сводной таблицы id болезни - id желаемого эффекта
    cursor.execute('''CREATE TABLE IF NOT EXISTS rel_deseases_effects 
                    (id_desease INTEGER,
                    id_effect Integer,
                    FOREIGN KEY (id_desease) REFERENCES spine_deseases(id)
                    FOREIGN KEY (id_effect) REFERENCES effects(id)
                    )''')
    for row in ws7.iter_rows(min_row=2, values_only=True):
        id_desease, id_effect_string = row
        id_effects = str(id_effect_string).split(', ')
        for id_effect in id_effects:
            cursor.execute('''INSERT INTO rel_deseases_effects (
                    id_desease,
                    id_effect
                    ) VALUES (?, ?)''', (id_desease, id_effect))   
    # Сохранение изменений и закрытие соединения с базой данных
    conn.commit()
    conn.close()
